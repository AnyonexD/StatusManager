# ============================================================
# StatusManager
# Layout original (Stack) + lógica completa de status/captura
# Organização: utilitários de caminho/credenciais unificados.
# ============================================================

import sys
import os
import time
import base64
import json
import shutil
import tempfile
import subprocess
import threading
import webbrowser
import urllib.request
from pathlib import Path
from datetime import datetime

import flet as ft
import pandas as pd
import numpy as np
import pyautogui as py

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, SessionNotCreatedException, WebDriverException

import chromedriver_autoinstaller
from dotenv import load_dotenv

from classes import *

# ============================================================
# CONSTANTES
# ============================================================

LOGIN_URL = "http://adm.desktop.com.br/login.jsp?"
ATIVOS_URL = "https://adm.desktop.com.br/Ativos.jsp"
MENU_URL = "https://adm.desktop.com.br/menu.jsp"
MFA_URL = "https://desktop.sso.e-trust.com.br/mfa/login/validate"
WHATSAPP_URL = "https://wa.me/5519920026971"
APP_VERSION = "2.1.1"
GITHUB_LATEST_RELEASE_API = (
    "https://api.github.com/repos/AnyonexD/StatusManager/releases/latest"
)
UPDATE_ASSET_NAME = "StatusManager.exe"

CHROME_UPDATE_MESSAGE = (
    "Não foi possível iniciar o Google Chrome automaticamente.\n\n"
    "Isso normalmente acontece quando o Google Chrome está desatualizado ou "
    "quando a versão do ChromeDriver não é compatível com o Chrome instalado.\n\n"
    "O que fazer:\n"
    "1. Abra o Google Chrome.\n"
    "2. Vá em Menu > Ajuda > Sobre o Google Chrome.\n"
    "3. Aguarde a atualização terminar.\n"
    "4. Feche e abra o Chrome novamente.\n"
    "5. Abra o StatusManager e tente outra vez."
)

# ============================================================
# PATH BASE (PyInstaller ou script normal)
# ============================================================

if hasattr(sys, "_MEIPASS"):
    BASE_DIR = Path(sys._MEIPASS)
else:
    BASE_DIR = Path(__file__).resolve().parent

ASSETS_DIR = os.path.join(str(BASE_DIR), "assets")

# ============================================================
# UTILITÁRIOS DE CAMINHO
# ============================================================

def _robos_candidates():
    user_home = Path.home()
    desktop_names = ("Desktop", "Área de Trabalho")
    robos_names = (".Robos", "Robos")
    candidates = []

    def add(path):
        if not path:
            return
        normalized = os.path.normpath(os.path.expandvars(str(path)))
        if normalized not in candidates:
            candidates.append(normalized)

    # Override para suporte/TI: ROBOS_HOME deve apontar para a pasta .Robos/Robos.
    add(os.getenv("ROBOS_HOME"))

    # Se o script ja esta dentro de Robos/.Robos, essa e a melhor referencia.
    try:
        for parent in Path(__file__).resolve().parents:
            if parent.name.lower() in ("robos", ".robos"):
                add(parent)
                break
    except Exception:
        pass

    # Variaveis do OneDrive mudam conforme conta corporativa/pessoal.
    for root in (os.getenv("OneDriveCommercial"), os.getenv("OneDriveConsumer"), os.getenv("OneDrive")):
        for desktop_name in desktop_names:
            for robos_name in robos_names:
                add(Path(root) / desktop_name / robos_name if root else None)

    # Variantes conhecidas usadas pela empresa/maquinas antigas.
    for one_drive_name in ("OneDrive - Desktop SA", "OneDrive - Desktop Sigmanet", "OneDrive"):
        for desktop_name in desktop_names:
            for robos_name in robos_names:
                add(user_home / one_drive_name / desktop_name / robos_name)

    # Desktop local, com ou sem OneDrive.
    for desktop_name in desktop_names:
        for robos_name in robos_names:
            add(user_home / desktop_name / robos_name)

    return candidates


def _can_use_robos_path(path):
    try:
        logins_dir = os.path.join(path, ".Logins")
        os.makedirs(logins_dir, exist_ok=True)
        test_path = os.path.join(logins_dir, ".write_test")
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test_path)
        return True
    except Exception:
        return False


def get_robos_base_path(logger=None):
    for path in _robos_candidates():
        if os.path.exists(path) and _can_use_robos_path(path):
            if logger:
                logger.info(f"Diretório Robos encontrado: {path}")
            return path
        elif os.path.exists(path) and logger:
            logger.warning(f"Diretório Robos encontrado, mas sem permissão de escrita: {path}")

    default_path = os.path.join(os.path.expanduser("~"), "Desktop", ".Robos")
    os.makedirs(default_path, exist_ok=True)
    if not _can_use_robos_path(default_path):
        raise PermissionError(f"Sem permissão de escrita na pasta Robos padrão: {default_path}")
    if logger:
        logger.info(f"Diretório Robos criado: {default_path}")
    return default_path


def get_logins_dir(base_path):
    logins_dir = os.path.join(base_path, ".Logins")
    os.makedirs(logins_dir, exist_ok=True)
    test_path = os.path.join(logins_dir, ".write_test")
    with open(test_path, "w", encoding="utf-8") as f:
        f.write("ok")
    os.remove(test_path)
    return logins_dir


def get_env_path(logger=None):
    env_path = os.path.join(get_logins_dir(get_robos_base_path(logger)), ".env")
    if logger:
        logger.info(f"Caminho de credenciais resolvido: {env_path}")
    return env_path


def get_validation_dir(base_path):
    validation_dir = os.path.join(base_path, "Code", "Alteração")
    os.makedirs(validation_dir, exist_ok=True)
    return validation_dir


def get_main_sheet_name(excel_path):
    """Retorna a primeira aba do arquivo, que é a aba principal do StatusManager."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    try:
        return wb.sheetnames[0] if wb.sheetnames else "Sheet1"
    finally:
        wb.close()


def salvar_tabela_principal(tabela, excel_path, sheet_name, logger=None):
    """Atualiza somente a aba principal, preservando as demais abas do Dados.xlsx."""
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        tabela.to_excel(writer, sheet_name=sheet_name, index=False)

    if logger:
        logger.info(f"Aba '{sheet_name}' atualizada sem remover as demais abas")

# ============================================================
# ATUALIZAÇÃO AUTOMÁTICA
# ============================================================

def _version_tuple(version):
    version = str(version).strip().lower().lstrip("v")
    parts = []
    for part in version.split("."):
        digits = "".join(ch for ch in part if ch.isdigit())
        parts.append(int(digits or 0))
    while len(parts) < 3:
        parts.append(0)
    return tuple(parts[:3])


def _request_json(url, timeout=12):
    req = urllib.request.Request(
        url,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": f"StatusManager/{APP_VERSION}",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _download_file(url, target_path, timeout=60):
    req = urllib.request.Request(url, headers={"User-Agent": f"StatusManager/{APP_VERSION}"})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        with open(target_path, "wb") as file:
            shutil.copyfileobj(response, file)


def is_chromedriver_version_error(error):
    text = str(error).lower()
    version_error_terms = (
        "session not created",
        "this version of chromedriver only supports",
        "only supports chrome version",
        "chrome version must be",
        "cannot find chrome binary",
        "chrome failed to start",
        "unable to obtain driver",
    )
    return any(term in text for term in version_error_terms)


def excel_value_to_text(value):
    if pd.isnull(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def buscar_release_mais_recente(logger=None):
    release = _request_json(GITHUB_LATEST_RELEASE_API)
    latest_version = str(release.get("tag_name") or "").lstrip("v")
    assets = release.get("assets") or []
    exe_asset = next((asset for asset in assets if asset.get("name") == UPDATE_ASSET_NAME), None)

    if not latest_version:
        raise ValueError("Release sem tag de versão")
    if not exe_asset:
        raise ValueError(f"Release sem o asset {UPDATE_ASSET_NAME}")

    info = {
        "version": latest_version,
        "download_url": exe_asset["browser_download_url"],
        "release_url": release.get("html_url", ""),
        "notes": release.get("body", ""),
    }
    if logger:
        logger.info(f"Release mais recente encontrada: v{latest_version}")
    return info


def iniciar_atualizador(downloaded_exe, logger=None):
    current_exe = sys.executable
    app_dir = os.path.dirname(current_exe)
    updater_path = os.path.join(tempfile.gettempdir(), "StatusManager_updater.bat")

    script = f"""@echo off
setlocal
set "TARGET={current_exe}"
set "NEW={downloaded_exe}"
timeout /t 2 /nobreak >nul
:trycopy
copy /Y "%NEW%" "%TARGET%" >nul
if errorlevel 1 (
    timeout /t 1 /nobreak >nul
    goto trycopy
)
start "" "%TARGET%"
del "%NEW%" >nul 2>nul
del "%~f0" >nul 2>nul
"""
    with open(updater_path, "w", encoding="utf-8") as file:
        file.write(script)

    if logger:
        logger.info(f"Atualizador criado em: {updater_path}")

    subprocess.Popen(
        ["cmd", "/c", updater_path],
        cwd=app_dir,
        creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
    )


def verificar_atualizacao_em_background(page, logger):
    def worker():
        try:
            release = buscar_release_mais_recente(logger)
            if _version_tuple(release["version"]) <= _version_tuple(APP_VERSION):
                logger.info(f"StatusManager já está atualizado: v{APP_VERSION}")
                return

            if not getattr(sys, "frozen", False):
                logger.info(
                    f"Nova versão disponível: v{release['version']}. "
                    "Atualização automática só roda no executável."
                )
                return

            resposta = py.confirm(
                text=(
                    f"Nova versão disponível: v{release['version']}\n"
                    f"Versão atual: v{APP_VERSION}\n\n"
                    "Deseja baixar e atualizar agora?"
                ),
                title="Atualização disponível",
                buttons=["Atualizar", "Depois"],
            )
            if resposta != "Atualizar":
                logger.info("Atualização recusada pelo usuário")
                return

            update_dir = os.path.join(tempfile.gettempdir(), "StatusManager_update")
            os.makedirs(update_dir, exist_ok=True)
            downloaded_exe = os.path.join(update_dir, UPDATE_ASSET_NAME)
            logger.info(f"Baixando atualização para: {downloaded_exe}")
            _download_file(release["download_url"], downloaded_exe)

            py.alert(
                "Atualização baixada com sucesso.\n\n"
                "O StatusManager será fechado, atualizado e aberto novamente.",
                title="Atualização pronta",
            )
            iniciar_atualizador(downloaded_exe, logger)
            page.window.close()
        except Exception as e:
            logger.error(f"Erro ao verificar atualização: {e}")

    threading.Thread(target=worker, daemon=True).start()

# ============================================================
# CREDENCIAIS
# ============================================================

def save_credentials(username, password, remember, logger):
    if not (remember and username and password):
        return False

    try:
        env_path = get_env_path(logger)
        temp_path = f"{env_path}.tmp"
        with open(temp_path, "w", encoding="utf-8") as f:
            f.write(f"LOGINADM={base64.b64encode(username.encode()).decode()}\n")
            f.write(f"SENHAADM={base64.b64encode(password.encode()).decode()}\n")
        os.replace(temp_path, env_path)
        logger.info(f"Credenciais salvas com sucesso em: {env_path}")
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar credenciais: {e}")
        return False


def load_saved_credentials(login_field, password_field, remember_checkbox, logger):
    try:
        env_path = get_env_path(logger)
        if not os.path.exists(env_path):
            logger.info(f"Nenhuma credencial salva encontrada em: {env_path}")
            return False

        load_dotenv(dotenv_path=env_path, override=True)
        login = os.getenv("LOGINADM")
        senha = os.getenv("SENHAADM")
        if login and senha:
            login_field.value = base64.b64decode(login).decode()
            password_field.value = base64.b64decode(senha).decode()
            remember_checkbox.value = True
            logger.info(f"Credenciais carregadas com sucesso de: {env_path}")
            return True
    except Exception as e:
        logger.error(f"Erro ao carregar credenciais: {e}")

    return False

# ============================================================
# HELPERS DE ALTERAÇÃO DE MODELO
# ============================================================

def _norm_modelo(texto):
    """Normaliza nome de modelo: remove '*', espaços extras e maiúsculas."""
    return str(texto).replace("*", "").strip().lower()


def _voltar_para_pesquisa(wait):
    """Volta da tela de edição para a tela de pesquisa, sem salvar."""
    try:
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/a[1]'))).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()
    except Exception:
        pass

# ============================================================
# HELPERS DE CONSULTA DE HISTÓRICO (modo "Consultar Histórico")
# ============================================================

def capturar_dados_ativo(chromedriver, wait, logger):
    """Lê os dados do ativo na tela 'Resultado da Pesquisa'.

    Retorna dict com codigo, mac_serial, sn_iccid e status.
    """
    base = '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr'
    dados = {"codigo": "", "mac_serial": "", "sn_iccid": "", "status": ""}
    try:
        dados["codigo"] = chromedriver.find_element(By.XPATH, f"{base}/td[1]").text.strip()

        # Célula MAC/Serial/ICCID: MAC em cima e "(s/n ou iccid: ...)" dentro de <small>
        celula_mac = chromedriver.find_element(By.XPATH, f"{base}/td[2]")
        texto_completo = celula_mac.text.strip()
        try:
            small = celula_mac.find_element(By.TAG_NAME, "small").text.strip()
        except Exception:
            small = ""
        dados["mac_serial"] = texto_completo.replace(small, "").strip()
        dados["sn_iccid"] = (small.replace("(s/n ou iccid:", "").replace(")", "").strip())

        dados["status"] = chromedriver.find_element(
            By.XPATH, f"{base}/td[3]").text.split("\n")[0].strip()

        logger.info(f"Dados do ativo capturados: {dados}")
        return dados
    except Exception as e:
        logger.error(f"Erro ao capturar dados do ativo: {e}")
        return dados


def capturar_ultima_log(chromedriver, wait, logger):
    """Abre o histórico do ativo e retorna a última (mais recente) entrada como dict:
    log_status, log_responsavel, log_evento, log_data.

    A tabela de histórico está em ordem cronológica crescente, então a última
    linha é a mais recente.
    """
    log = {"log_status": "", "log_responsavel": "", "log_evento": "", "log_data": ""}
    try:
        # Abre o histórico do ativo (ícone a[2] na coluna de Ações)
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[4]/a[2]/img'))).click()
        logger.info("Histórico aberto")

        # A tabela de histórico é a 3ª da página
        wait.until(EC.presence_of_element_located((By.XPATH, '(//table)[3]/tbody/tr')))
        linhas = chromedriver.find_elements(By.XPATH, '(//table)[3]/tbody/tr')

        # Remove cabeçalho e linhas vazias
        linhas_dados = [l for l in linhas[1:] if l.text.strip()]
        if not linhas_dados:
            logger.info("Histórico sem registros")
            return log

        # Mais recente = última linha
        ultima = linhas_dados[-1]
        celulas = ultima.find_elements(By.TAG_NAME, "td")

        # Célula 1: status (linha 1) + responsável (linha 2)
        if len(celulas) > 0:
            partes = [p.strip() for p in celulas[0].text.split("\n") if p.strip()]
            log["log_status"] = partes[0] if len(partes) > 0 else ""
            log["log_responsavel"] = partes[1] if len(partes) > 1 else ""

        # Célula 2: evento + observações (junta tudo numa linha)
        if len(celulas) > 1:
            log["log_evento"] = celulas[1].text.strip().replace("\n", " ")

        # Célula 3: data
        if len(celulas) > 2:
            log["log_data"] = celulas[2].text.strip()

        logger.info(f"Última log capturada: {log}")
        return log
    except Exception as e:
        logger.error(f"Erro ao capturar histórico: {e}")
        return log


def capturar_historico_completo(chromedriver, wait, logger):
    """Abre o histórico do ativo e retorna TODAS as logs como lista de dicts.

    Cada item tem: log_status, log_responsavel, log_evento, log_data.
    A ordem segue a da tabela (cronológica crescente: mais antiga primeiro).
    """
    historico = []
    try:
        # Abre o histórico do ativo (ícone a[2] na coluna de Ações)
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[4]/a[2]/img'))).click()
        logger.info("Histórico aberto (completo)")

        # A tabela de histórico é a 3ª da página
        wait.until(EC.presence_of_element_located((By.XPATH, '(//table)[3]/tbody/tr')))
        linhas = chromedriver.find_elements(By.XPATH, '(//table)[3]/tbody/tr')

        # Remove cabeçalho e linhas vazias
        linhas_dados = [l for l in linhas[1:] if l.text.strip()]
        if not linhas_dados:
            logger.info("Histórico sem registros")
            return historico

        for linha in linhas_dados:
            celulas = linha.find_elements(By.TAG_NAME, "td")
            log = {"log_status": "", "log_responsavel": "", "log_evento": "", "log_data": ""}

            if len(celulas) > 0:
                partes = [p.strip() for p in celulas[0].text.split("\n") if p.strip()]
                log["log_status"] = partes[0] if len(partes) > 0 else ""
                log["log_responsavel"] = partes[1] if len(partes) > 1 else ""
            if len(celulas) > 1:
                log["log_evento"] = celulas[1].text.strip().replace("\n", " ")
            if len(celulas) > 2:
                log["log_data"] = celulas[2].text.strip()

            historico.append(log)

        logger.info(f"Histórico completo capturado: {len(historico)} registros")
        return historico
    except Exception as e:
        logger.error(f"Erro ao capturar histórico completo: {e}")
        return historico


def gravar_detalhes(excel_path, serie, dados, log, logger):
    """Adiciona uma linha na aba 'Detalhes' do Dados.xlsx com dados do ativo + última log."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        if "Detalhes" not in wb.sheetnames:
            ws = wb.create_sheet("Detalhes")
            ws.append([
                "Ativo/Mac/Serie", "Código (Ativo)", "MAC/Serial/ICCID", "S/N ou ICCID",
                "Status Atual", "Log - Status", "Log - Responsável",
                "Log - Evento/Observações", "Log - Data", "Horário Processado",
            ])
        else:
            ws = wb["Detalhes"]

        ws.append([
            serie,
            dados.get("codigo", ""),
            dados.get("mac_serial", ""),
            dados.get("sn_iccid", ""),
            dados.get("status", ""),
            log.get("log_status", ""),
            log.get("log_responsavel", ""),
            log.get("log_evento", ""),
            log.get("log_data", ""),
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        ])
        wb.save(excel_path)
        logger.info(f"Linha gravada na aba Detalhes para: {serie}")
    except Exception as e:
        logger.error(f"Erro ao gravar na aba Detalhes: {e}")


def gravar_detalhes_completo(excel_path, serie, dados, historico, logger):
    """Adiciona VÁRIAS linhas na aba 'Detalhes' (uma por log do histórico completo),
    repetindo os dados do ativo em cada linha."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        if "Detalhes" not in wb.sheetnames:
            ws = wb.create_sheet("Detalhes")
            ws.append([
                "Ativo/Mac/Serie", "Código (Ativo)", "MAC/Serial/ICCID", "S/N ou ICCID",
                "Status Atual", "Log - Status", "Log - Responsável",
                "Log - Evento/Observações", "Log - Data", "Horário Processado",
            ])
        else:
            ws = wb["Detalhes"]

        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Se não houver histórico, grava ao menos uma linha com os dados do ativo
        if not historico:
            ws.append([
                serie, dados.get("codigo", ""), dados.get("mac_serial", ""),
                dados.get("sn_iccid", ""), dados.get("status", ""),
                "", "", "(sem registros no histórico)", "", agora,
            ])
        else:
            for log in historico:
                ws.append([
                    serie,
                    dados.get("codigo", ""),
                    dados.get("mac_serial", ""),
                    dados.get("sn_iccid", ""),
                    dados.get("status", ""),
                    log.get("log_status", ""),
                    log.get("log_responsavel", ""),
                    log.get("log_evento", ""),
                    log.get("log_data", ""),
                    agora,
                ])

        wb.save(excel_path)
        logger.info(f"{len(historico) or 1} linha(s) gravada(s) na aba Detalhes para: {serie}")
    except Exception as e:
        logger.error(f"Erro ao gravar histórico completo na aba Detalhes: {e}")

# ============================================================
# ROBÔ DE CADASTRO DE ATIVOS (modo "Cadastrar Ativos")
# ============================================================

def processar_cadastro(chromedriver, wait, excel_path, update_progress, logger):
    """Lê a aba 'Cadastrar' do Dados.xlsx e cadastra cada ativo no site.

    Mantém a lógica original do robô CadastrarAtivos: reaproveita a tela quando
    o modelo é o mesmo do item anterior, trata SIMCARD separadamente e grava o
    resultado (Cadastrado / mensagem de erro) na coluna 'Situacao'.

    O login já foi feito pelo StatusManager; aqui só navegamos para a tela de
    cadastro e processamos a planilha.
    """
    import pandas as pd

    # Lê a aba Cadastrar
    try:
        tabela = pd.read_excel(excel_path, sheet_name="Cadastrar")
        logger.info(f"Aba 'Cadastrar' lida: {len(tabela)} linhas")
    except Exception as e:
        logger.error(f"Erro ao ler aba 'Cadastrar': {e}")
        py.alert(f"Erro ao ler a aba 'Cadastrar': {e}")
        return

    if "Situacao" not in tabela.columns:
        tabela["Situacao"] = ""
    tabela["Situacao"] = tabela["Situacao"].astype("object")

    def salvar():
        # Salva preservando as outras abas do arquivo
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            tabela.to_excel(writer, sheet_name="Cadastrar", index=False)

    # Navega para a tela de cadastro de ativos
    chromedriver.get(ATIVOS_URL)
    time.sleep(0.5)
    wait.until(EC.presence_of_element_located(
        (By.XPATH, '/html/body/table[2]/tbody/tr[8]/td/input'))).click()

    modelo_anterior = None
    total = len(tabela)
    sucesso = 0
    erro = 0
    update_progress(0, total)

    for i, row in tabela.iterrows():
        id_ = excel_value_to_text(row.get("ID", ""))
        modelo = excel_value_to_text(row.get("Modelo", "")).upper()
        ativo = excel_value_to_text(row.get("Ativo", ""))
        mac = excel_value_to_text(row.get("Mac", ""))
        sn = excel_value_to_text(row.get("Serie", ""))
        fsan = excel_value_to_text(row.get("Fsan", ""))
        observacao = excel_value_to_text(row.get("Observacao", ""))
        nota_fiscal = excel_value_to_text(row.get("Nota Fiscal", ""))

        logger.info(f"Cadastro {i + 1}/{total}: {modelo} - {ativo}")

        try:
            # --- Seleção do modelo (com reaproveitamento de tela) ---
            if modelo_anterior is None:
                element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipoSelect"]')))
                Select(element).select_by_value(str(id_))
                wait.until(EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()

            elif modelo_anterior == modelo and modelo == "SIMCARD":
                try:
                    if chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b').text == "A operação solicitada não pôde ser concluída":
                        chromedriver.execute_script("window.history.go(-2)")
                        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipoSelect"]')))
                        element = chromedriver.find_element(By.XPATH, '//*[@id="tipoSelect"]')
                        Select(element).select_by_value(str(id_))
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()
                except Exception:
                    element = chromedriver.find_element(By.XPATH, '//*[@id="tipoSelect"]')
                    Select(element).select_by_value(str(id_))
                    wait.until(EC.presence_of_element_located(
                        (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()

            elif modelo_anterior == modelo:
                try:
                    if chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b').text == "A operação solicitada não pôde ser concluída":
                        chromedriver.execute_script("window.history.go(-2)")
                        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipoSelect"]')))
                        element = chromedriver.find_element(By.XPATH, '//*[@id="tipoSelect"]')
                        Select(element).select_by_value(str(id_))
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()
                except NoSuchElementException:
                    chromedriver.execute_script("window.history.go(-2)")
                    wait.until(EC.presence_of_element_located(
                        (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()

            elif modelo_anterior != modelo:
                wait.until(EC.presence_of_element_located(
                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b')))
                try:
                    if chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b').text == "A operação solicitada não pôde ser concluída":
                        chromedriver.execute_script("window.history.go(-2)")
                        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipoSelect"]')))
                        element = chromedriver.find_element(By.XPATH, '//*[@id="tipoSelect"]')
                        Select(element).select_by_value(str(id_))
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()
                except Exception:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipoSelect"]')))
                    element = chromedriver.find_element(By.XPATH, '//*[@id="tipoSelect"]')
                    Select(element).select_by_value(str(id_))
                    wait.until(EC.presence_of_element_located(
                        (By.XPATH, '/html/body/div[2]/form/div[1]/div[2]/button'))).click()

            # --- Confirma modelo carregado ---
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div/h6/font/b')))
            modelo_atual = chromedriver.find_element(
                By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div/h6/font/b').text

            # --- CASO SIMCARD ---
            if modelo == modelo_atual.upper() and modelo == "SIMCARD":
                campo = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="codigo"]')))
                campo.clear(); campo.send_keys(ativo)
                campo = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="serial"]')))
                campo.clear(); campo.send_keys(sn)
                time.sleep(2)

                if observacao:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="observacoes"]')))
                    chromedriver.find_element(By.XPATH, '//*[@id="observacoes"]').clear()
                    chromedriver.find_element(By.XPATH, '//*[@id="observacoes"]').send_keys(observacao)

                wait.until(EC.presence_of_element_located(
                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div/div[2]/table/tbody/tr/td/button'))).click()

                try:
                    if chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b').text == "A operação solicitada não pôde ser concluída":
                        errotxt = chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div/div[1]/p').text
                        tabela.loc[i, "Situacao"] = errotxt
                        erro += 1
                    else:
                        tabela.loc[i, "Situacao"] = "Cadastrado"
                        sucesso += 1
                except Exception:
                    tabela.loc[i, "Situacao"] = "Cadastrado"
                    sucesso += 1

                modelo_anterior = modelo

            # --- CASO outros modelos ---
            elif modelo == modelo_atual.upper():
                if not fsan:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="codigo"]')))
                    chromedriver.find_element(By.XPATH, '//*[@id="codigo"]').clear()
                    chromedriver.find_element(By.XPATH, '//*[@id="mac"]').clear()
                    chromedriver.find_element(By.XPATH, '//*[@id="serial"]').clear()
                    chromedriver.find_element(By.XPATH, '//*[@id="info"]').clear()

                    chromedriver.find_element(By.XPATH, '//*[@id="codigo"]').send_keys(ativo)
                    time.sleep(0.3)
                    chromedriver.find_element(By.XPATH, '//*[@id="mac"]').send_keys(mac)
                    time.sleep(0.3)
                    chromedriver.find_element(By.XPATH, '//*[@id="serial"]').send_keys(sn)
                    time.sleep(0.3)

                    if observacao:
                        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="observacoes"]')))
                        chromedriver.find_element(By.XPATH, '//*[@id="observacoes"]').clear()
                        chromedriver.find_element(By.XPATH, '//*[@id="observacoes"]').send_keys(observacao)
                else:
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="codigo"]'))).clear()
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mac"]'))).clear()
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="serial"]'))).clear()
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="info"]'))).clear()

                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="codigo"]'))).send_keys(ativo)
                    time.sleep(0.3)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mac"]'))).send_keys(mac)
                    time.sleep(0.3)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="serial"]'))).send_keys(sn)
                    time.sleep(0.3)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="info"]'))).send_keys(fsan)
                    time.sleep(0.3)

                    if observacao:
                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="observacoes"]'))).clear()
                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="observacoes"]'))).send_keys(observacao)

                if nota_fiscal:
                    chromedriver.find_element(By.XPATH, '//*[@id="nota"]').clear()
                    chromedriver.find_element(By.XPATH, '//*[@id="nota"]').send_keys(nota_fiscal)

                wait.until(EC.presence_of_element_located(
                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div/div[2]/table/tbody/tr/td/button'))).click()
                time.sleep(1)

                try:
                    if chromedriver.find_element(
                            By.XPATH, '/html/body/table[2]/tbody/tr/td/div/h4/b').text == "A operação solicitada não pôde ser concluída":
                        errotxt = wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div/div[1]/p'))).text
                        tabela.loc[i, "Situacao"] = errotxt
                        erro += 1
                    else:
                        tabela.loc[i, "Situacao"] = "Cadastrado"
                        sucesso += 1
                except Exception:
                    tabela.loc[i, "Situacao"] = "Cadastrado"
                    sucesso += 1

                modelo_anterior = modelo

        except Exception as e:
            logger.error(f"Erro ao cadastrar item {i + 1}: {e}")
            tabela.loc[i, "Situacao"] = f"Erro: {e}"
            erro += 1

        salvar()
        update_progress(i + 1, total)

    logger.info(f"Cadastro finalizado. Sucessos: {sucesso}, Erros: {erro}")
    py.alert(f"Cadastro finalizado!\n\nProcessados: {total}\nSucessos: {sucesso}\nErros: {erro}")

# ============================================================
# APP PRINCIPAL
# ============================================================

def main(page: ft.Page):
    # ---- Janela ----
    page.window.width = page.window.max_width = page.window.min_width = 375
    page.window.height = page.window.max_height = page.window.min_height = 500
    page.window.center()
    page.window.title_bar_hidden = True
    page.adaptive, page.padding = True, 0
    page.window.icon = os.path.join(ASSETS_DIR, "desk_logo.ico")
    page.title = f"StatusManager v{APP_VERSION}"

    logger = Logger_Manager("robos_log.log")

    # ---- Componentes base ----
    login_field = Login_and_pass.login()
    password_field = Login_and_pass.password()
    status_dropdown = Status_Selection.dropdown()

    remember_checkbox = ft.Checkbox(
        label="Lembrar Usuário e Senha",
        label_style=ft.TextStyle(size=12, color=ft.Colors.WHITE),
        value=False,
    )
    try:
        env_display_path = get_env_path(logger)
    except Exception as e:
        env_display_path = f"Erro ao resolver .env: {e}"
        logger.error(env_display_path)

    def compact_path(path):
        path = str(path)
        drive = os.path.splitdrive(path)[0]
        marker = f"{os.sep}Robos{os.sep}"
        marker_hidden = f"{os.sep}.Robos{os.sep}"

        if marker in path:
            return f"{drive}{os.sep}...{path[path.index(marker):]}"
        if marker_hidden in path:
            return f"{drive}{os.sep}...{path[path.index(marker_hidden):]}"
        if ".Logins" in path:
            return f"{drive}{os.sep}...{path[path.index('.Logins'):]}"
        return f"{drive}{os.sep}...{os.path.basename(path)}" if drive else os.path.basename(path)

    env_path_text = ft.Text(
        value=f"Credenciais: {compact_path(env_display_path)}",
        color=ft.Colors.WHITE70,
        size=9,
        max_lines=1,
        overflow=ft.TextOverflow.ELLIPSIS,
        tooltip=env_display_path,
    )

    def copy_env_path(e):
        page.set_clipboard(env_display_path)
        logger.info(f"Caminho do .env copiado: {env_display_path}")
        page.snack_bar = ft.SnackBar(ft.Text("Caminho do .env copiado"))
        page.snack_bar.open = True
        page.update()

    progress_bar = ft.ProgressBar(
        width=220, color=ft.Colors.YELLOW_700, bgcolor="#555555", value=0
    )
    progress_counter = ft.Text(value="0/0", color=ft.Colors.WHITE, size=12)
    info_text = ft.Text(value="Progresso:", color=ft.Colors.WHITE, size=12)

    def update_progress(current, total):
        progress_counter.value = f"{current}/{total}"
        progress_bar.value = (current / total) if total else 0
        page.update()

    # ========================================================
    # AÇÕES DE UI (pasta, whatsapp, fechar)
    # ========================================================

    def open_whatsapp(e):
        webbrowser.open(WHATSAPP_URL)
        logger.info("Link do WhatsApp aberto")

    def open_project_folder(e):
        base_path = get_robos_base_path(logger)
        folder = os.path.join(base_path, "Code", "Alteração")
        os.makedirs(folder, exist_ok=True)
        os.startfile(folder)
        logger.info(f"Pasta do projeto aberta: {folder}")

    # ========================================================
    # FLUXO PRINCIPAL DE AUTOMAÇÃO (lógica original completa)
    # ========================================================

    def process_login():
        username = excel_value_to_text(login_field.value)
        password = excel_value_to_text(password_field.value)
        selected_status = status_dropdown.value

        if not username or not password:
            py.alert("Informe login e senha antes de iniciar.", title="Credenciais obrigatórias")
            restaurar_botao()
            return

        if not selected_status:
            py.alert("Erro: É necessário escolher uma opção: 'Disponível', "
                     "'Defeito/Inutilizado', 'Baixado/Perdido' ou 'Procurar'")
            restaurar_botao()
            return

        if selected_status == "Alterar Modelo":
            py.alert("A opção 'Alterar Modelo' foi removida desta versão.")
            restaurar_botao()
            return

        base_path = get_robos_base_path(logger)
        validation_dir = get_validation_dir(base_path)
        excel_path = os.path.join(validation_dir, "Dados.xlsx")
        logger.info(f"Caminho do Excel: {excel_path}")

        # Cria planilha se não existir
        if not os.path.exists(excel_path):
            try:
                pd.DataFrame(columns=[
                    "Ativo/Mac/Serie", "Tipo", "Horario Processado",
                    "Ativo", "Antigo Status", "Observação"
                ]).to_excel(excel_path, sheet_name="Status", index=False)
                logger.info(f"Arquivo Excel criado: {excel_path}")
            except Exception as e:
                logger.error(f"Erro ao criar arquivo Excel: {e}")
                py.alert(f"Erro ao criar o arquivo Excel: {e}")
                restaurar_botao()
                return
        else:
            # Verifica se o arquivo está aberto
            try:
                with open(excel_path, "r+b"):
                    pass
            except IOError:
                py.alert("Erro: O arquivo 'Dados.xlsx' está aberto.\n"
                         "Por favor, salve e feche o arquivo antes de continuar.")
                restaurar_botao()
                return

        save_credentials(username, password, remember_checkbox.value, logger)

        try:
            logger.info("Iniciando processo de automação com Chrome")
            try:
                chromedriver_autoinstaller.install()
                options = webdriver.ChromeOptions()
                options.add_experimental_option("detach", True)
                options.add_argument("--start-maximized")
                options.add_argument("--allow-running-insecure-content")
                options.add_argument("--allow-insecure-localhost")
                chromedriver = webdriver.Chrome(options=options)
            except (SessionNotCreatedException, WebDriverException) as e:
                logger.error(f"Erro ao iniciar Chrome/ChromeDriver: {e}")
                if is_chromedriver_version_error(e):
                    py.alert(CHROME_UPDATE_MESSAGE, title="Atualize o Google Chrome")
                else:
                    py.alert(f"Erro ao iniciar o Google Chrome:\n\n{e}", title="Erro no Chrome")
                return

            wait = WebDriverWait(chromedriver, 30)

            login_decoded = username
            senha_decoded = password

            os.chdir(validation_dir)

            try:
                main_sheet_name = get_main_sheet_name(excel_path)
                tabela = pd.read_excel(excel_path, sheet_name=main_sheet_name)
                logger.info(f"Arquivo Excel lido: {len(tabela)} linhas")

                # Garante que colunas de texto aceitem strings (evita erro float64)
                colunas_texto = [
                    "Modelo", "Modelo Novo", "ID", "Ativo", "Mac", "N.Série",
                    "FSAN", "Antigo Status", "Novo Status", "Observação",
                    "Horario Processado",
                ]
                for col in colunas_texto:
                    if col in tabela.columns:
                        tabela[col] = tabela[col].astype("object")
            except Exception as e:
                logger.error(f"Erro ao ler arquivo Excel: {e}")
                py.alert(f"Erro ao ler o arquivo Excel: {e}")
                chromedriver.quit()
                return

            total_rows = len(tabela)
            current_row = 0
            update_progress(current_row, total_rows)

            # ---- LOGIN ----
            chromedriver.get(LOGIN_URL)
            logger.info("Acessando o SSO")
            wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'input[type="text"][name="username"]'))).send_keys(login_decoded)
            wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'input[type="password"]'))).send_keys(senha_decoded)
            wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button[type="submit"]'))).click()

            logger.info("Aguardando o input do usuário...")
            if not chromedriver.current_url.startswith(MENU_URL):
                wait.until(EC.url_matches(MFA_URL))

            try:
                wait = WebDriverWait(chromedriver, 120)
                wait.until(EC.url_matches(MENU_URL))
                wait = WebDriverWait(chromedriver, 30)
            except Exception as e:
                logger.info(f"Houve um erro durante o login: {e}")
                chromedriver.quit()
                return

            logger.info("Login realizado, aguardando carregamento da página")

            # --- MODO CADASTRAR ATIVOS: fluxo próprio, lê da aba "Cadastrar" ---
            if selected_status == "Cadastrar Ativos":
                processar_cadastro(chromedriver, wait, excel_path, update_progress, logger)
                progress_bar.color = ft.Colors.GREEN
                page.update()
                py.alert("Após confirmar, você será redirecionado para o arquivo!")
                try:
                    chromedriver.close()
                except Exception:
                    pass
                os.startfile(excel_path)
                return

            chromedriver.get(ATIVOS_URL)
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/table[2]/tbody/tr[6]/td/table/tbody/tr/td/font/input[2]"))).click()
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/table[2]/tbody/tr[8]/td/input"))).click()
            logger.info("Navegação inicial concluída, iniciando processamento")

            total_rows = len(tabela[tabela["Horario Processado"] != "nan"])
            current_row = 0

            # ---- LOOP DE ITENS ----
            for i, serie in enumerate(tabela["Ativo/Mac/Serie"]):
                if tabela.at[i, "Horario Processado"] == "nan":
                    continue

                serie = excel_value_to_text(serie)
                tipo = excel_value_to_text(tabela.at[i, "Tipo"]) if "Tipo" in tabela.columns else ""
                tipo = tipo.upper()
                tabela.at[i, "Horario Processado"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                logger.info(f"Processando item {i + 1}/{total_rows}: {serie}, Tipo: {tipo}")

                if not serie:
                    logger.info(f"Item {i + 1} sem série, pulando")
                    tabela.at[i, "Antigo Status"] = "Linha sem Ativo/Mac/Serie preenchido"
                    continue

                try:
                    # --- Preenche chave de busca ---
                    campo_chave = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="chave"]')))
                    campo_chave.clear()
                    campo_chave.send_keys(serie)

                    # --- Seleciona o critério de busca conforme o tipo ---
                    if tipo == "ATIVO":
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div/div[2]/select/option[1]'))).click()
                    elif tipo == "MAC":
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div/div[2]/select/option[2]'))).click()
                    elif tipo in ("SERIE", "SÉRIE"):
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '//html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div/div[2]/select/option[3]'))).click()
                    elif tipo == "FSAN":
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div/div[2]/select/option[4]'))).click()

                    wait.until(EC.presence_of_element_located(
                        (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[2]/button'))).click()

                    # --- Verifica se nada foi encontrado ---
                    try:
                        nao_encontrado = wait.until(EC.presence_of_element_located(
                            (By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/p"))).text
                        if nao_encontrado == "Total de ativos encontrados: 0":
                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()
                            tabela.at[i, "Antigo Status"] = (
                                f"Nenhum equipamento localizado utilizando o critério de busca: {tipo}.")
                            salvar_tabela_principal(tabela, excel_path, main_sheet_name, logger)
                            logger.info(f"Item {i + 1} não encontrado")
                            current_row += 1
                            update_progress(current_row, total_rows)
                            continue
                    except Exception:
                        pass

                    try:
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[3]')))
                        status_atual = chromedriver.find_element(
                            By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[3]").text
                        status_atual = status_atual.split("\n")[0]
                        logger.info(f"Status atual do item {i + 1}: {status_atual}")

                        # --- CASO 1: Alocado (e não é modo Procurar) ---
                        if "Alocado" in status_atual and selected_status not in ("Procurar", "Histórico (Última Log)", "Histórico (Completo)"):
                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]/b')))
                            numero_ativo = chromedriver.find_element(
                                By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]/b").text

                            tabela.at[i, "Ativo"] = numero_ativo
                            tabela.at[i, "Antigo Status"] = status_atual
                            tabela.at[i, "Novo Status"] = status_atual
                            salvar_tabela_principal(tabela, excel_path, main_sheet_name, logger)
                            logger.info(f"Item {i + 1} está alocado, pulando")

                            wait.until(EC.element_to_be_clickable(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()
                            current_row += 1
                            update_progress(current_row, total_rows)
                            continue

                        # --- CASO 2: Modo Procurar (captura dados, não altera) ---
                        try:
                            if selected_status == "Procurar":
                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[4]/a[1]/img'))).click()

                                ativo = wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[1]/div[2]/input'))).get_attribute("value")
                                tabela.at[i, "Ativo"] = ativo

                                mac = wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[1]/div[3]/input[1]'))).get_attribute("value")
                                tabela.at[i, "Mac"] = mac

                                nserie = wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[2]/div[1]/input[1]'))).get_attribute("value")
                                tabela.at[i, "N.Série"] = nserie

                                fsan = wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[2]/div[2]/input[1]'))).get_attribute("value")
                                tabela.at[i, "FSAN"] = fsan

                                modelo_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tipo"]')))
                                modelo = Select(modelo_element)
                                wait.until(lambda d: len(modelo.options) > 0)
                                modelo_texto = modelo.first_selected_option.text
                                tabela.at[i, "Modelo"] = modelo_texto.replace("*", "").strip()

                                tabela.at[i, "Antigo Status"] = status_atual
                                tabela.at[i, "Novo Status"] = status_atual

                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/a[1]'))).click()
                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]/img'))).click()

                                salvar_tabela_principal(tabela, excel_path, main_sheet_name, logger)
                                current_row += 1
                                update_progress(current_row, total_rows)
                                continue
                        except Exception:
                            wait.until(EC.element_to_be_clickable(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/a[1]'))).click()
                            wait.until(EC.element_to_be_clickable(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()
                            continue

                        # --- CASO Histórico: captura dados do ativo + log(s) ---
                        if selected_status in ("Histórico (Última Log)", "Histórico (Completo)"):
                            # 1. Captura os dados do ativo na tela de resultado
                            dados = capturar_dados_ativo(chromedriver, wait, logger)

                            # 2. Captura a(s) log(s) conforme o modo escolhido
                            if selected_status == "Histórico (Completo)":
                                historico = capturar_historico_completo(chromedriver, wait, logger)
                                gravar_detalhes_completo(excel_path, serie, dados, historico, logger)
                            else:
                                log = capturar_ultima_log(chromedriver, wait, logger)
                                gravar_detalhes(excel_path, serie, dados, log, logger)

                            # 3. Volta para a tela de pesquisa para o próximo item
                            try:
                                chromedriver.get(ATIVOS_URL)
                                wait.until(EC.presence_of_element_located(
                                    (By.XPATH, "/html/body/table[2]/tbody/tr[6]/td/table/tbody/tr/td/font/input[2]"))).click()
                                wait.until(EC.presence_of_element_located(
                                    (By.XPATH, "/html/body/table[2]/tbody/tr[8]/td/input"))).click()
                            except Exception:
                                _voltar_para_pesquisa(wait)

                            current_row += 1
                            update_progress(current_row, total_rows)
                            continue

                        # --- CASO 3: Status já é o desejado (apenas observação) ---
                        if ((selected_status == "Disponível" and "Livre" in status_atual) or
                                (selected_status == "Defeito/Inutilizado" and "Defeito" in status_atual)):

                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]')))
                            numero_ativo = chromedriver.find_element(
                                By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]").text

                            tabela.at[i, "Ativo"] = numero_ativo
                            tabela.at[i, "Antigo Status"] = status_atual
                            obs = tabela.at[i, "Observação"] if "Observação" in tabela.columns else ""
                            obs = np.nan if str(obs).lower() == "nan" else obs

                            if pd.notna(obs) and obs.strip() != "":
                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[4]/a[1]/img'))).click()
                                campo_obs = wait.until(EC.presence_of_element_located(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[3]/div[1]/input[1]')))
                                campo_obs.clear()
                                campo_obs.send_keys(obs)
                                wait.until(EC.presence_of_element_located(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[2]/table/tbody/tr/td/button'))).click()
                                time.sleep(2)
                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()
                            else:
                                wait.until(EC.element_to_be_clickable(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()

                            tabela.at[i, "Novo Status"] = status_atual
                            salvar_tabela_principal(tabela, excel_path, main_sheet_name, logger)
                            current_row += 1
                            update_progress(current_row, total_rows)
                            continue

                        # --- CASO 4: Precisa alterar o status ---
                        else:
                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]/b')))
                            numero_ativo = chromedriver.find_element(
                                By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[1]/b").text

                            tabela.at[i, "Ativo"] = numero_ativo
                            tabela.at[i, "Antigo Status"] = status_atual
                            obs = tabela.at[i, "Observação"] if "Observação" in tabela.columns else ""
                            obs = np.nan if str(obs).lower() == "nan" else obs

                            wait.until(EC.element_to_be_clickable(
                                (By.XPATH, '//html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[4]/a[1]/img'))).click()

                            status_value = "L"
                            if selected_status == "Defeito/Inutilizado":
                                status_value = "D"
                            if selected_status == "Baixado/Perdido":
                                status_value = "B"

                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[2]/div[3]/select'))).send_keys(status_value)

                            if not (pd.isnull(obs) or str(obs).strip().lower() in ["nan", "null", "none", ""]):
                                campo_obs = wait.until(EC.presence_of_element_located(
                                    (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[1]/div[3]/div[1]/input[1]')))
                                campo_obs.clear()
                                campo_obs.send_keys(obs)

                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/form/div/div[2]/table/tbody/tr/td/button'))).click()
                            time.sleep(2)

                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[3]')))
                            status_new = chromedriver.find_element(
                                By.XPATH, "/html/body/table[2]/tbody/tr/td/div/div[1]/div/div/table/tbody/tr/td[3]").text
                            tabela.at[i, "Novo Status"] = status_new

                            wait.until(EC.presence_of_element_located(
                                (By.XPATH, '/html/body/table[2]/tbody/tr/td/div/div[2]/a[1]'))).click()

                        salvar_tabela_principal(tabela, excel_path, main_sheet_name, logger)
                        current_row += 1
                        update_progress(current_row, total_rows)

                    except Exception:
                        wait.until(EC.element_to_be_clickable(
                            (By.XPATH, '/html/body/table[2]/tbody/tr[8]/td/a[1]/img')))
                        chromedriver.find_element(
                            By.XPATH, "/html/body/table[2]/tbody/tr[8]/td/a[1]/img").click()
                        current_row += 1
                        update_progress(current_row, total_rows)
                        continue

                except Exception:
                    try:
                        wait.until(EC.element_to_be_clickable(
                            (By.XPATH, '/html/body/table[2]/tbody/tr[8]/td/a[1]/img')))
                        chromedriver.find_element(
                            By.XPATH, "/html/body/table[2]/tbody/tr[8]/td/a[1]/img").click()
                    except Exception:
                        pass
                    current_row += 1
                    update_progress(current_row, total_rows)
                    continue

            open_xlsx = os.path.join(base_path, "Code", "Alteração", "Dados.xlsx")
            # Feedback de conclusão: barra fica verde
            progress_bar.color = ft.Colors.GREEN
            page.update()
            py.alert("Após confirmar, você será redirecionado para o arquivo!")
            chromedriver.close()
            os.startfile(open_xlsx)

        except Exception as e:
            py.alert(f"Erro ao executar o script: {e}")
            try:
                chromedriver.quit()
            except Exception:
                pass
        finally:
            # Restaura o botão e a cor da barra
            restaurar_botao()
            progress_bar.color = ft.Colors.YELLOW_700
            page.update()

    # ========================================================
    # COMPONENTES VISUAIS (layout original)
    # ========================================================

    botao_pasta = ft.IconButton(
        icon=ft.Icons.FOLDER,
        icon_color=ft.Colors.WHITE,
        tooltip="Clique aqui para abrir a pasta aonde contem o arquivo excel.",
        on_click=open_project_folder,
    )

    whatsapp_button = ft.Container(
        content=ft.IconButton(
            icon=ft.Icons.CONTACT_SUPPORT,
            tooltip="Clique aqui para tirar dúvidas ou ver mais informações.",
            icon_color=ft.Colors.WHITE,
            icon_size=20,
            on_click=open_whatsapp,
        ),
        width=42,
        height=42,
        top=430,
        left=315,
    )

    status_pasta_row = ft.Row(
        controls=[status_dropdown, botao_pasta],
        alignment=ft.MainAxisAlignment.START,
        spacing=20,
        width=290,
    )
    status_pasta_container = ft.Container(
        content=status_pasta_row,
        margin=ft.margin.only(left=30, top=320),
    )

    # ---- Botão de login com animação ----
    def login_function(e):
        # Estado "processando": muda texto e destaca
        login_button.content.controls[0].name = ft.Icons.HOURGLASS_TOP
        login_button.content.controls[1].value = "Processando..."
        login_button.bgcolor = ft.Colors.YELLOW_700
        login_button.content.controls[0].color = ft.Colors.BLACK
        login_button.content.controls[1].color = ft.Colors.BLACK
        login_button.border = ft.border.all(2, ft.Colors.YELLOW_700)
        page.update()

        threading.Thread(target=process_login, daemon=True).start()

    def restaurar_botao():
        login_button.content.controls[0].name = ft.Icons.ARROW_FORWARD
        login_button.content.controls[1].value = "Entrar"
        login_button.bgcolor = ft.Colors.TRANSPARENT
        login_button.content.controls[0].color = ft.Colors.YELLOW_700
        login_button.content.controls[1].color = ft.Colors.YELLOW_700
        login_button.border = ft.border.all(2, ft.Colors.YELLOW_700)
        page.update()

    login_button = ft.Container(
        content=ft.Row([
            ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.YELLOW_700, size=18),
            ft.Text("Entrar", size=16, color=ft.Colors.YELLOW_700, weight=ft.FontWeight.W_500),
        ], alignment=ft.MainAxisAlignment.CENTER, spacing=10),
        width=300,
        height=45,
        bgcolor=ft.Colors.TRANSPARENT,
        border=ft.border.all(2, ft.Colors.YELLOW_700),
        border_radius=8,
        alignment=ft.Alignment(0, 0),
        animate=ft.Animation(250, ft.AnimationCurve.EASE_IN_OUT),
        on_click=login_function,
    )

    # ---- Progresso ----
    progress_bar_container = ft.Container(content=progress_bar, margin=ft.margin.only(top=5))
    progress_column = ft.Column(
        controls=[
            ft.Row([info_text, progress_counter], spacing=5,
                   alignment=ft.MainAxisAlignment.START),
            progress_bar_container,
        ],
        spacing=0,
        horizontal_alignment=ft.CrossAxisAlignment.START,
    )
    progress_counter_container = ft.Container(
        content=progress_column,
        margin=ft.margin.only(left=30, top=380),
    )

    login_button_container = ft.Container(
        content=login_button,
        margin=ft.margin.only(left=18, top=430),
    )

    # ---- Botão fechar ----
    def hover_close(e):
        close.content.opacity = 1 if e.data == "true" else 0
        close.content.update()

    close = ft.Container(
        margin=ft.margin.only(left=335),
        content=ft.Icon(
            ft.Icons.CLOSE,
            color=ft.Colors.WHITE,
            opacity=0,
            animate_opacity=ft.Animation(duration=600, curve=ft.AnimationCurve.EASE_IN_OUT),
        ),
        ink=True,
        on_hover=hover_close,
        on_click=lambda _: page.window.close(),
    )

    # ---- Stack principal ----
    logo_container = ft.Container(
        content=ft.Image(
            src=os.path.join(ASSETS_DIR, "Images", "logo.png"),
            width=250, height=75, fit=ft.ImageFit.CONTAIN,
        ),
        margin=ft.margin.only(top=25, left=47),
        opacity=0,
        animate_opacity=ft.Animation(500, ft.AnimationCurve.EASE_OUT),
    )
    login_box = ft.Container(
        width=280, height=50, margin=ft.margin.only(left=30, top=125), content=login_field,
        opacity=0, animate_opacity=ft.Animation(500, ft.AnimationCurve.EASE_OUT),
    )
    password_box = ft.Container(
        width=280, height=50, margin=ft.margin.only(left=30, top=185), content=password_field,
        opacity=0, animate_opacity=ft.Animation(500, ft.AnimationCurve.EASE_OUT),
    )
    remember_box = ft.Container(
        margin=ft.margin.only(left=30, top=245), content=remember_checkbox,
        opacity=0, animate_opacity=ft.Animation(500, ft.AnimationCurve.EASE_OUT),
    )
    env_path_box = ft.Container(
        width=315,
        margin=ft.margin.only(left=30, top=282),
        content=ft.Row(
            controls=[
                env_path_text,
                ft.IconButton(
                    icon=ft.Icons.CONTENT_COPY,
                    icon_color=ft.Colors.WHITE70,
                    icon_size=14,
                    tooltip="Copiar caminho do .env",
                    on_click=copy_env_path,
                ),
            ],
            spacing=4,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        opacity=0,
        animate_opacity=ft.Animation(500, ft.AnimationCurve.EASE_OUT),
    )

    stack_main = ft.Stack(
        expand=True,
        controls=[
            ft.Container(
                content=ft.Image(
                    src=os.path.join(ASSETS_DIR, "Images", "bgdesktop.png"),
                    width=400, height=550, fit=ft.ImageFit.COVER,
                )
            ),
            ft.Container(
                content=ft.Text(
                    f"© 2026 Desktop | Desenvolvido por Ygor Guedes | v{APP_VERSION}",
                    color=ft.Colors.WHITE, size=12, text_align=ft.TextAlign.CENTER,
                ),
                width=375, height=20, left=0, bottom=5,
                alignment=ft.Alignment(0, 0),
            ),
            logo_container,
            ft.WindowDragArea(
                width=375, height=100,
                content=ft.Container(bgcolor="transparent"),
            ),
            close,
            login_box,
            password_box,
            remember_box,
            env_path_box,
            status_pasta_container,
            progress_counter_container,
            login_button_container,
            whatsapp_button,
        ],
    )

    # ---- Carrega credenciais e monta a página ----
    if load_saved_credentials(login_field, password_field, remember_checkbox, logger):
        page.update()

    page.add(stack_main)
    verificar_atualizacao_em_background(page, logger)

    # ---- Fade-in de entrada (em sequência) ----
    def _fade_in():
        for box in (logo_container, login_box, password_box, remember_box, env_path_box):
            box.opacity = 1
            page.update()
            time.sleep(0.12)

    threading.Thread(target=_fade_in, daemon=True).start()


if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")
